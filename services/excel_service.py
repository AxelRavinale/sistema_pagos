"""
Servicio para generar archivos Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class ExcelService:
    
    @staticmethod
    def generar_planilla(planilla_id):
        """
        Genera el archivo Excel de una planilla.
        
        Args:
            planilla_id (int): ID de la planilla
            
        Returns:
            str: Ruta del archivo generado
        """
        # 1. Obtener datos de la planilla
        from models.planilla import Planilla
        planilla = Planilla.obtener_por_id(planilla_id)
        items = planilla.obtener_items()
        
        # 2. Asignar números de cheque (AQUÍ es cuando se asignan)
        items_cheques = [i for i in items if i['modalidad_pago'] in [6, 8]]
        if items_cheques:
            from services.cheque_service import ChequeService
            asignaciones = ChequeService.asignar_numeros_a_planilla(
                planilla_id, 
                items_cheques
            )
            
            # Actualizar los items con los números asignados
            for item in items_cheques:
                item['numero_cheque'] = asignaciones[item['id']]
        
        # 3. Crear el Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilla de Pagos"
        
        # 4. Agregar encabezados (fila 3 según tu plantilla)
        encabezados = [
            "Tipo de documento",
            "Número de documento",
            "Sucursal",
            "Identificación del pago",
            "Denominación del beneficiario",
            "Importe",
            "Cuenta de débito",
            "Cuenta de pago - CBU o Nº Cheque",
            "Modalidad de Pago",
            "Marca de registración de cheque",
            "Fecha de pago - Emisión",
            "Fecha de pago diferido"
        ]
        
        for col, header in enumerate(encabezados, start=2):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # 5. Agregar datos
        fila = 4
        for item in items:
            ws.cell(fila, 2, item['tipo_documento'])
            ws.cell(fila, 3, item['numero_documento'])
            ws.cell(fila, 4, planilla.sucursal)
            ws.cell(fila, 5, item['identificacion_pago'])
            ws.cell(fila, 6, item['beneficiario'])
            ws.cell(fila, 7, item['importe'])
            ws.cell(fila, 8, planilla.cuenta_debito)
            
            # CBU o número de cheque
            if item['modalidad_pago'] in [6, 8]:
                ws.cell(fila, 9, item['numero_cheque'])
            else:
                ws.cell(fila, 9, item['cuenta_pago'])
            
            ws.cell(fila, 10, item['modalidad_pago'])
            ws.cell(fila, 11, item.get('marca_registracion', ''))
            ws.cell(fila, 12, item.get('fecha_emision', ''))
            ws.cell(fila, 13, item.get('fecha_pago_diferido', ''))
            
            fila += 1
        
        # 6. Ajustar anchos de columna
        for col in range(2, 14):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # 7. Guardar archivo
        filename = f"planilla_{planilla.referencia}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"/home/claude/planillas_generadas/{filename}"
        
        wb.save(filepath)
        
        # 8. Actualizar estado de la planilla
        planilla.marcar_como_generada(filepath)
        
        return filepath